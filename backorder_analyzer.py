#!/usr/bin/env python3
"""
Navision Backorder Analyzer
===========================

Dit script analyseert Navision backorder exports en genereert een overzichtelijke Excel
met verzendbare en backorder artikelen, inclusief automatische categorisering en e-mailgeneratie.

Auteur: AUTONAV
Versie: 2.0
Datum: 2024
"""

import pandas as pd
import numpy as np
import logging
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Import configuratie
try:
    from config import *
except ImportError:
    # Fallback configuratie als config.py niet bestaat
    INPUT_FILE = "Navision_Backorder_Export_Realistic.xlsx"
    OUTPUT_FILE = "Output/Backorder_Analyse.xlsx"
    LOCATION_CODE = "DSV"
    FULLY_RESERVED = "No"
    ORDER_STATUS = "Backorder"
    COLORS = {
        'header': '366092',
        'sendable': 'C6EFCE',
        'backorder': 'FFC7CE',
        'order_header': 'D9E1F2',
        'border': '000000',
        'category_1': 'FF6B6B',
        'category_2': '4ECDC4',
        'category_3': 'FFA500',
        'category_4': '9B59B6'
    }
    COLUMN_WIDTHS = [15, 40, 12, 12, 15, 40, 12, 12, 20]
    LOG_FILE = "backorder_analyzer.log"
    LOG_LEVEL = "INFO"
    REQUIRED_COLUMNS = [
        'Sales Order No.', 'Item No.', 'Description', 'Quantity',
        'Quantity Available', 'Location Code', 'Fully Reserved',
        'Order Status', 'Customer Name'
    ]
    # Categorie configuratie
    CATEGORY_1_ITEMS = ['OIL-5W30-1L', 'OIL-10W40-1L', 'BRAKE-FLUID', 'COOLANT-1L']
    CATEGORY_2_ITEMS = ['BRAKE-FLUID-DOT4', 'BRAKE-PADS-FRONT', 'BRAKE-PADS-REAR']
    CATEGORY_3_ITEMS = ['EXHAUST-CAT', 'EXHAUST-MUFFLER', 'EXHAUST-PIPE']
    CATEGORIES = {
        1: {'name': 'Bestel bij fabrikant', 'color': 'FF6B6B'},
        2: {'name': 'Binnenkort leverbaar', 'color': '4ECDC4'},
        3: {'name': 'Geen voorraadvooruitzicht', 'color': 'FFA500'}
    }
    EMAIL_TEMPLATES = {}
    SALESFORCE_EMAIL_SETTINGS = {'enabled': False}

# Import CategoryManager
try:
    from category_manager import CategoryManager
    category_manager = CategoryManager()
except ImportError:
    category_manager = None

# Setup logging
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL),
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def load_navision_data(file_path):
    """Laad de Navision export data."""
    logging.info(f"Laden van Navision export: {file_path}")
    
    try:
        df = pd.read_excel(file_path)
        logging.info(f"Data geladen: {len(df)} rijen, {len(df.columns)} kolommen")
        
        # Debug: Check voor negatieve quantities
        if 'QUANTITY' in df.columns:
            negative_quantities = df[df['QUANTITY'] < 0]
            if len(negative_quantities) > 0:
                logging.warning(f"⚠️  GEVONDEN: {len(negative_quantities)} rijen met negatieve QUANTITY waarden!")
                logging.warning(f"Voorbeelden: {negative_quantities[['DOCUMENT_ID', 'TYPE_ID', 'QUANTITY', 'AVAILABLE_STOCK']].head().to_dict('records')}")
        
        return df
    except Exception as e:
        logging.error(f"Fout bij laden van bestand: {e}")
        raise

def validate_columns(df):
    """Valideer en map kolommen naar het verwachte formaat."""
    logging.info(f"Beschikbare kolommen: {list(df.columns)}")
    
    # Kolom mapping voor jouw Excel formaat
    column_mapping = {
        'DOCUMENT_ID': 'Sales Order No.',
        'SELL_TO_CUSTOMER_ID': 'Customer Name', 
        'TYPE_ID': 'Item No.',
        'QUANTITY': 'Quantity',
        'AVAILABLE_STOCK': 'Quantity Available'
    }
    
    # Verwijder ITEM_ID kolom als die bestaat (omdat we TYPE_ID gebruiken)
    if 'ITEM_ID' in df.columns:
        df = df.drop(columns=['ITEM_ID'])
        logging.info("ITEM_ID kolom verwijderd (gebruik TYPE_ID)")
    
    # Hernoem kolommen volgens mapping
    for old_col, new_col in column_mapping.items():
        if old_col in df.columns:
            df = df.rename(columns={old_col: new_col})
            logging.info(f"Kolom hernoemd: {old_col} -> {new_col}")
    
    # Voeg ontbrekende kolommen toe met default waarden (na hernoeming)
    df['Description'] = 'Artikel ' + df['Item No.'].astype(str)
    df['Location Code'] = 'DSV'  # Default locatie
    df['Fully Reserved'] = 'No'  # Default waarde
    df['Order Status'] = 'Backorder'  # Default status
    
    # Controleer of alle verplichte kolommen nu aanwezig zijn
    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        logging.warning(f"Ontbrekende kolommen na mapping: {missing_columns}")
        # Voeg ontbrekende kolommen toe met default waarden
        for col in missing_columns:
            if col == 'Description':
                df[col] = 'Artikel ' + df['Item No.'].astype(str)
            elif col == 'Location Code':
                df[col] = 'DSV'
            elif col == 'Fully Reserved':
                df[col] = 'No'
            elif col == 'Order Status':
                df[col] = 'Backorder'
            else:
                df[col] = ''
    
    logging.info("Kolom mapping voltooid")
    return df

def filter_backorder_data(df):
    """Filter de data op basis van de criteria."""
    original_count = len(df)
    
    # Filter op Location Code (alleen als niet leeg)
    if LOCATION_CODE:
        df = df[df['Location Code'] == LOCATION_CODE]
        logging.info(f"Na Location Code filter ({LOCATION_CODE}): {len(df)} rijen")
    
    # Filter op Fully Reserved (alleen als niet leeg)
    if FULLY_RESERVED:
        df = df[df['Fully Reserved'] == FULLY_RESERVED]
        logging.info(f"Na Fully Reserved filter ({FULLY_RESERVED}): {len(df)} rijen")
    
    # Filter op Order Status (alleen als niet leeg)
    if ORDER_STATUS:
        df = df[df['Order Status'] == ORDER_STATUS]
        logging.info(f"Na Order Status filter ({ORDER_STATUS}): {len(df)} rijen")
    
    logging.info(f"Filtering voltooid: {original_count} -> {len(df)} rijen")
    return df

def categorize_backorder_items(df):
    """Categoriseer backorder artikelen in de drie categorieën."""
    def get_category(item_no):
        if category_manager:
            category = category_manager.get_category_for_item(item_no)
            # Als CategoryManager geen categorie heeft (None), gebruik dan geen categorie
            if category is None:
                return None
            return category
        else:
            # Geen CategoryManager beschikbaar - alle artikelen krijgen geen categorie
            return None
    
    # Maak een kopie van de DataFrame om pandas warnings te voorkomen
    df_copy = df.copy()
    
    # Voeg categorie toe aan DataFrame
    df_copy['Category'] = df_copy['Item No.'].apply(get_category)
    
    # Bepaal categorie namen en acties
    df_copy['Category_Name'] = df_copy['Category'].apply(lambda x: 
        category_manager.get_category_name(x) if category_manager 
        else 'Geen categorie'
    )
    
    df_copy['Category_Action'] = df_copy['Category'].apply(lambda x: 
        category_manager.get_category_action(x) if category_manager 
        else 'Behoud backorder'
    )
    
    return df_copy

def group_by_sales_order(df):
    """Groepeer data per Sales Order en categoriseer artikelen."""
    grouped_data = {}
    
    for order_no in df['Sales Order No.'].unique():
        order_data = df[df['Sales Order No.'] == order_no]
        customer_name = order_data['Customer Name'].iloc[0]
        
        # Split in verzendbaar en backorder
        # Gebruik pandas.isna() om NaN waarden te controleren
        sendable = order_data[(order_data['Quantity Available'] > 0) & (~order_data['Quantity Available'].isna())]
        # Backorder: 0, negatieve waarden, of NaN
        backorder = order_data[(order_data['Quantity Available'] <= 0) | (order_data['Quantity Available'].isna())]
        
        # Check voor BA/HP artikelen (batterijen/fietsen) - deze mogen gewoon als backorder blijven
        # Converteer naar string en check veilig
        ba_hp_items = backorder[backorder['Item No.'].astype(str).str.startswith(('BA', 'HP'), na=False)]
        if len(ba_hp_items) > 0:
            logging.info(f"Order {order_no} bevat {len(ba_hp_items)} BA/HP artikelen (batterijen/fietsen) - deze blijven als normale backorder")
        
        # Als er 0 verzendbaar en 0 backorder zijn, check dan voor BA/HP nummers
        if len(sendable) == 0 and len(backorder) == 0:
            # Check of er BA/HP nummers in de hele order zitten
            all_ba_hp = order_data[order_data['Item No.'].astype(str).str.startswith(('BA', 'HP'), na=False)]
            if len(all_ba_hp) > 0:
                logging.info(f"Order {order_no} bevat alleen BA/HP artikelen (fiets/batterij) - niet aanpassen")
                # Markeer als speciale categorie voor Excel output
                grouped_data[order_no] = {
                    'customer': customer_name,
                    'sendable': sendable,
                    'backorder': backorder,
                    'total_items': len(order_data),
                    'sendable_count': len(sendable),
                    'backorder_count': len(backorder),
                    'is_bike_battery': True
                }
                continue
        
        # Debug: Check voor onmogelijke situaties
        if len(sendable) == 0 and len(backorder) == 0:
            logging.warning(f"Order {order_no} heeft geen verzendbare of backorder artikelen!")
            logging.warning(f"Order data: {order_data[['Item No.', 'Quantity', 'Quantity Available']].to_dict('records')}")
            # Forceer alle artikelen als backorder als er geen verzendbare of backorder zijn
            backorder = order_data.copy()
            logging.info(f"Order {order_no}: Alle artikelen geforceerd als backorder")
        
        # Categoriseer backorder artikelen
        if len(backorder) > 0:
            backorder = categorize_backorder_items(backorder)
        
        grouped_data[order_no] = {
            'customer': customer_name,
            'sendable': sendable,
            'backorder': backorder,
            'total_items': len(order_data),
            'sendable_count': len(sendable),
            'backorder_count': len(backorder)
        }
        
        logging.info(f"Order {order_no} ({customer_name}): {len(sendable)} verzendbaar, {len(backorder)} backorder")
    
    return grouped_data

def shorten_url(url):
    """Verkort een URL door alleen het domein en belangrijke delen te behouden."""
    if not url or not isinstance(url, str):
        return url
    
    # Verwijder protocol
    if url.startswith('http://'):
        url = url[7:]
    elif url.startswith('https://'):
        url = url[8:]
    
    # Verwijder www. als aanwezig
    if url.startswith('www.'):
        url = url[4:]
    
    # Als de URL nog steeds te lang is, behoud alleen het domein en eerste pad
    if len(url) > 50:
        parts = url.split('/')
        if len(parts) > 1:
            # Behoud domein + eerste pad
            shortened = '/'.join(parts[:2])
            if len(shortened) > 50:
                # Als nog steeds te lang, behoud alleen domein
                shortened = parts[0]
            return shortened
        else:
            return parts[0]
    
    return url

def generate_email_content(item_data, category):
    """Genereer e-mail content voor een artikel."""
    if category not in EMAIL_TEMPLATES:
        return None
    
    template = EMAIL_TEMPLATES[category]
    
    # Vul template variabelen
    email_data = {
        'customer_name': item_data['Customer Name'],
        'item_no': item_data['Item No.'],
        'item_description': item_data.get('Description', f'Artikel {item_data["Item No."]}'),
        'order_no': item_data['Sales Order No.'],
        'quantity': item_data['Quantity']
    }
    
    # Voeg specifieke links toe via CategoryManager
    if category_manager:
        if category == 1:
            # Categorie 1: Link naar fabrikant
            manufacturer_link = category_manager.get_item_link(item_data['Item No.'], 'fabrikant')
            if manufacturer_link:
                email_data['manufacturer_message'] = f"Wij raden u aan om dit artikel direct bij de fabrikant te bestellen:\n🔗 {shorten_url(manufacturer_link)} (verkorte link)"
            else:
                email_data['manufacturer_message'] = "Wij raden u aan om contact op te nemen met de fabrikant van dit artikel voor bestelling."
        elif category == 3:
            # Categorie 3: Link naar externe verkoper
            external_link = category_manager.get_item_link(item_data['Item No.'], 'externe_verkoper')
            if external_link:
                email_data['external_seller_message'] = f"Wij raden u aan om dit artikel bij een externe verkoper te bestellen:\n🔗 {shorten_url(external_link)} (verkorte link)"
            else:
                email_data['external_seller_message'] = "Wij raden u aan om een externe verkoper te zoeken voor dit artikel."
    else:
        # Fallback naar oude methode
        if category == 1:
            manufacturer_links = template.get('manufacturer_links', {})
            link = manufacturer_links.get(
                item_data['Item No.'], 
                manufacturer_links.get('default', 'https://www.original-equipment-parts.com')
            )
            email_data['manufacturer_message'] = f"Wij raden u aan om dit artikel direct bij de fabrikant te bestellen:\n🔗 {shorten_url(link)} (verkorte link)"
        elif category == 3:
            external_links = template.get('external_seller_links', {})
            link = external_links.get(
                item_data['Item No.'], 
                external_links.get('default', 'https://www.autodoc.nl')
            )
            email_data['external_seller_message'] = f"Wij raden u aan om dit artikel bij een externe verkoper te bestellen:\n🔗 {shorten_url(link)} (verkorte link)"
    
    # Genereer e-mail content
    subject = template['subject'].format(**email_data)
    body = template['body'].format(**email_data)
    
    return {
        'to': item_data['Customer Name'],
        'subject': subject,
        'body': body.strip(),
        'category': category,
        'item_data': item_data
    }

def create_excel_workbook(grouped_data):
    """Maak een Excel werkboek met de geanalyseerde data."""
    logging.info("Excel werkboek succesvol aangemaakt")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Backorder Analyse"
    
    # Definieer kleuren
    colors = {
        'header': PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid'),
        'sendable': PatternFill(start_color=COLORS['sendable'], end_color=COLORS['sendable'], fill_type='solid'),
        'backorder': PatternFill(start_color=COLORS['backorder'], end_color=COLORS['backorder'], fill_type='solid'),
        'order_header': PatternFill(start_color=COLORS['order_header'], end_color=COLORS['order_header'], fill_type='solid'),
        'category_1': PatternFill(start_color=COLORS['category_1'], end_color=COLORS['category_1'], fill_type='solid'),
        'category_2': PatternFill(start_color=COLORS['category_2'], end_color=COLORS['category_2'], fill_type='solid'),
        'category_3': PatternFill(start_color=COLORS['category_3'], end_color=COLORS['category_3'], fill_type='solid'),
        'category_4': PatternFill(start_color=COLORS['category_4'], end_color=COLORS['category_4'], fill_type='solid'),
        'no_category': PatternFill(start_color=COLORS['backorder'], end_color=COLORS['backorder'], fill_type='solid')
    }
    
    # Definieer borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Definieer fonts
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font()
    
    current_row = 1
    
    # Voor elke order
    for order_no, order_info in grouped_data.items():
        customer = order_info['customer']
        
        # Check voor fiets/batterij orders
        if order_info.get('is_bike_battery', False):
            ws.cell(row=current_row, column=1, value=f"Order: {order_no}")
            ws.cell(row=current_row, column=2, value=f"Klant: {customer}")
            ws.cell(row=current_row, column=3, value="🚲 FIETS/BATTERIJ ORDER")
            ws.cell(row=current_row, column=4, value="NIET AANPASSEN")
            ws.cell(row=current_row, column=5, value="")
            
            # Styling voor fiets/batterij order
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type='solid')  # Goud
                cell.font = Font(bold=True)
                cell.border = thin_border
        else:
            # Normale order header
            ws.cell(row=current_row, column=1, value=f"Order: {order_no}")
            ws.cell(row=current_row, column=2, value=f"Klant: {customer}")
            ws.cell(row=current_row, column=3, value=f"Totaal: {order_info['total_items']}")
            ws.cell(row=current_row, column=4, value=f"Verzendbaar: {order_info['sendable_count']}")
            ws.cell(row=current_row, column=5, value=f"Backorder: {order_info['backorder_count']}")
            
            # Styling voor normale order header
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = colors['order_header']
                cell.font = Font(bold=True)
                cell.border = thin_border
        
        # Styling voor order header
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = colors['order_header']
            cell.font = Font(bold=True)
            cell.border = thin_border
        
        current_row += 1
        
        # Skip verdere verwerking voor fiets/batterij orders
        if order_info.get('is_bike_battery', False):
            current_row += 1  # Extra ruimte
            continue
        
        # Verzendbare artikelen
        if len(order_info['sendable']) > 0:
            ws.cell(row=current_row, column=1, value="✅ VERZENDBAAR")
            ws.cell(row=current_row, column=1).fill = colors['sendable']
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            # Headers voor verzendbare artikelen
            headers = ['Artikelnummer', 'Omschrijving', 'Aantal', 'Beschikbaar']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = colors['header']
                cell.font = header_font
                cell.border = thin_border
            current_row += 1
            
            # Verzendbare artikelen data
            for _, item in order_info['sendable'].iterrows():
                ws.cell(row=current_row, column=1, value=item['Item No.'])
                # Fallback voor Description als deze niet beschikbaar is
                description = item.get('Description', f'Artikel {item["Item No."]}')
                ws.cell(row=current_row, column=2, value=description)
                ws.cell(row=current_row, column=3, value=item['Quantity'])
                ws.cell(row=current_row, column=4, value=item['Quantity Available'])
                
                # Styling
                for col in range(1, 5):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = colors['sendable']
                    cell.border = thin_border
                current_row += 1
        
        # Backorder artikelen
        if len(order_info['backorder']) > 0:
            ws.cell(row=current_row, column=1, value="❌ BACKORDER")
            ws.cell(row=current_row, column=1).fill = colors['backorder']
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            # Headers voor backorder artikelen
            headers = ['Artikelnummer', 'Omschrijving', 'Aantal', 'Beschikbaar', 'Categorie', 'Actie']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.fill = colors['header']
                cell.font = header_font
                cell.border = thin_border
            current_row += 1
            
            # Backorder artikelen data
            for _, item in order_info['backorder'].iterrows():
                category = item['Category']
                category_name = item['Category_Name']
                
                ws.cell(row=current_row, column=1, value=item['Item No.'])
                # Fallback voor Description als deze niet beschikbaar is
                description = item.get('Description', f'Artikel {item["Item No."]}')
                ws.cell(row=current_row, column=2, value=description)
                ws.cell(row=current_row, column=3, value=item['Quantity'])
                ws.cell(row=current_row, column=4, value=item['Quantity Available'])
                ws.cell(row=current_row, column=5, value=category_name)
                
                # Actie kolom - gebruik de actie uit de data
                action = item.get('Category_Action', 'Behoud backorder')
                ws.cell(row=current_row, column=6, value=action)
                
                # Styling met categorie kleur
                if category is not None and not pd.isna(category):
                    category_color = colors[f'category_{int(category)}']
                else:
                    # Default kleur voor artikelen zonder categorie
                    category_color = colors['no_category']
                
                for col in range(1, 7):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = category_color
                    cell.border = thin_border
                current_row += 1
        
        # Spacing tussen orders
        current_row += 2
    
    # Auto-fit kolommen (veilige versie)
    for col in range(1, 7):  # Max 6 kolommen
        max_length = 0
        column_letter = ws.cell(row=1, column=col).column_letter
        
        for row in range(1, current_row):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def generate_email_report(grouped_data):
    """Genereer een rapport van alle e-mails die verzonden moeten worden."""
    email_report = []
    
    for order_no, order_info in grouped_data.items():
        for _, item in order_info['backorder'].iterrows():
            category = item['Category']
            
            # Alleen categorie 1 en 3 krijgen e-mails
            if category in [1, 3]:
                email_content = generate_email_content(item, category)
                if email_content:
                    email_report.append(email_content)
    
    return email_report

def save_excel_file(wb, file_path):
    """Sla het Excel bestand op."""
    # Maak output directory als deze niet bestaat
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    wb.save(file_path)
    logging.info(f"Excel bestand opgeslagen: {file_path}")

def save_email_report(email_report, file_path):
    """Sla het e-mail rapport op als Excel bestand."""
    if not email_report:
        logging.info("Geen e-mails om te verzenden")
        return
    
    # Maak DataFrame van e-mails
    email_data = []
    for email in email_report:
        email_data.append({
            'Ordernummer': email['item_data']['Sales Order No.'],
            'Klant': email['to'],
            'Artikelnummer': email['item_data']['Item No.'],
            'Omschrijving': email['item_data'].get('Description', f'Artikel {email["item_data"]["Item No."]}'),
            'Aantal': email['item_data']['Quantity'],
            'Categorie': email['category'],
            'Onderwerp': email['subject'],
            'E-mail Body': email['body']
        })
    
    df = pd.DataFrame(email_data)
    
    # Sla op als Excel
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='E-mails', index=False)
        
        # Auto-fit kolommen
        worksheet = writer.sheets['E-mails']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    logging.info(f"E-mail rapport opgeslagen: {file_path}")

def main(input_file=None):
    """Hoofdfunctie van het script."""
    logging.info("=== Navision Backorder Analyzer gestart ===")
    
    # Gebruik opgegeven file of fallback naar config
    file_to_use = input_file if input_file else INPUT_FILE
    logging.info(f"Gebruik bestand: {file_to_use}")
    
    try:
        # Laad data
        df = load_navision_data(file_to_use)
        
        # Valideer kolommen
        validate_columns(df)
        
        # Filter data
        filtered_df = filter_backorder_data(df)
        
        # Zorg ervoor dat de kolom namen correct zijn na filtering
        if 'DOCUMENT_ID' in filtered_df.columns:
            filtered_df = filtered_df.rename(columns={
                'DOCUMENT_ID': 'Sales Order No.',
                'SELL_TO_CUSTOMER_ID': 'Customer Name',
                'TYPE_ID': 'Item No.',
                'QUANTITY': 'Quantity',
                'AVAILABLE_STOCK': 'Quantity Available'
            })
            logging.info("Kolommen opnieuw hernoemd na filtering")
        
        # Groepeer per order
        grouped_data = group_by_sales_order(filtered_df)
        
        # Maak Excel werkboek
        wb = create_excel_workbook(grouped_data)
        
        # Genereer unieke bestandsnaam
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"Output/Backorder_Analyse_v{timestamp}.xlsx"
        
        # Sla Excel op
        save_excel_file(wb, output_file)
        
        # Genereer e-mail rapport
        email_report = generate_email_report(grouped_data)
        if email_report:
            email_file = OUTPUT_FILE.replace('.xlsx', '_Emails.xlsx')
            save_email_report(email_report, email_file)
        
        # Print samenvatting
        total_orders = len(grouped_data)
        total_sendable = sum(order['sendable_count'] for order in grouped_data.values())
        total_backorder = sum(order['backorder_count'] for order in grouped_data.values())
        
        logging.info("=== Analyse voltooid ===")
        logging.info(f"Totaal orders: {total_orders}")
        logging.info(f"Totaal verzendbare artikelen: {total_sendable}")
        logging.info(f"Totaal backorder artikelen: {total_backorder}")
        logging.info(f"Output bestand: {OUTPUT_FILE}")
        
        if email_report:
            logging.info(f"E-mails om te verzenden: {len(email_report)}")
            logging.info(f"E-mail rapport: {email_file}")
        
        # Categorie statistieken
        category_counts = {}
        for order_info in grouped_data.values():
            for _, item in order_info['backorder'].iterrows():
                category = item['Category']
                if category not in category_counts:
                    category_counts[category] = 0
                category_counts[category] += 1
        
        logging.info("Backorder categorieën:")
        for cat_num, count in category_counts.items():
            if count > 0:
                if cat_num is None:
                    logging.info(f"  - Geen categorie: {count} artikelen")
                elif category_manager:
                    cat_name = category_manager.get_category_name(cat_num)
                    logging.info(f"  - {cat_name}: {count} artikelen")
                else:
                    logging.info(f"  - Categorie {cat_num}: {count} artikelen")
        
    except Exception as e:
        import traceback
        logging.error(f"Fout tijdens uitvoering: {e}")
        logging.error(f"Traceback: {traceback.format_exc()}")
        raise

if __name__ == "__main__":
    main()